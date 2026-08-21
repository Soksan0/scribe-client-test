from __future__ import annotations

import math
import tempfile
from pathlib import Path
from typing import Any

from .profiling import profile_delimited


def load_frame(path: Path, file_format: str, profile: dict[str, Any] | None = None) -> Any:
    import pandas as pd

    profile = profile or {}
    parsing = profile.get("parsing_config") or {}
    if file_format in {"csv", "tsv"}:
        encoding = parsing.get("encoding") or profile.get("encoding") or "utf-8"
        delimiter = parsing.get("delimiter") or profile.get("delimiter") or ("\t" if file_format == "tsv" else ",")
        return pd.read_csv(path, dtype=object, keep_default_na=False, encoding=encoding, sep=delimiter, header=int(parsing.get("header_row") or 1) - 1)
    if file_format == "xlsx":
        sheet = profile.get("table_name") or profile.get("format_metadata", {}).get("profiled_sheet") or 0
        return pd.read_excel(path, sheet_name=sheet, dtype=object, keep_default_na=False, header=int(parsing.get("header_row") or 1) - 1)
    if file_format == "sav":
        import pyreadstat
        return pyreadstat.read_sav(path, apply_value_formats=False)[0]
    if file_format == "dta":
        import pyreadstat
        return pyreadstat.read_dta(path, apply_value_formats=False)[0]
    if file_format == "rds":
        import pyreadr
        objects = pyreadr.read_r(path)
        if not objects:
            raise ValueError("The RDS file does not contain a readable tabular object.")
        return next(iter(objects.values()))
    raise ValueError(f"Unsupported dataset format: {file_format}")


def preview_records(path: Path, file_format: str, profile: dict[str, Any] | None, offset: int, limit: int) -> dict[str, Any]:
    frame = load_frame(path, file_format, profile)
    subset = frame.iloc[offset : offset + limit]
    rows = []
    for index, record in subset.iterrows():
        values = {str(column): _plain_value(value) for column, value in record.items()}
        rows.append({"row_number": int(frame.index.get_loc(index)) + 1, "values": values})
    return {"columns": [str(column) for column in frame.columns], "rows": rows, "total": len(frame), "offset": offset, "limit": limit}


def _plain_value(value: Any) -> str:
    if value is None:
        return ""
    try:
        if math.isnan(value):
            return ""
    except (TypeError, ValueError):
        pass
    return str(value)


def _profile_frame(frame: Any, filename: str, warnings: list[str] | None = None) -> dict[str, Any]:
    with tempfile.NamedTemporaryFile(suffix=".csv") as handle:
        frame.to_csv(handle.name, index=False)
        result = profile_delimited(Path(handle.name), Path(filename).stem + ".csv")
    result["warnings"] = [*(warnings or []), *result["warnings"]]
    return result


def profile_xlsx(path: Path, filename: str, parsing_config: dict[str, Any] | None = None) -> dict[str, Any]:
    import pandas as pd
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=False)
    visible = [sheet for sheet in workbook.worksheets if sheet.sheet_state == "visible"]
    if not visible:
        raise ValueError("The workbook has no visible worksheets.")
    sheet = next((candidate for candidate in visible if candidate.max_row > 1 and candidate.max_column > 0), visible[0])
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"The worksheet {sheet.title!r} is empty.")
    header_row = int((parsing_config or {}).get("header_row") or 1)
    if header_row < 1 or header_row > len(rows):
        raise ValueError(f"Header row {header_row} is outside worksheet {sheet.title!r}")
    headers = [_plain_value(value).strip() or f"unnamed_{index + 1}" for index, value in enumerate(rows[header_row - 1])]
    frame = pd.DataFrame([[_plain_value(value) for value in row] for row in rows[header_row:]], columns=headers)
    formula_count = sum(isinstance(value, str) and value.startswith("=") for row in rows[header_row:] for value in row)
    warnings = []
    if len(visible) > 1:
        warnings.append(f"Profiled worksheet {sheet.title!r}; {len(visible) - 1} other visible worksheet(s) remain unchanged.")
    if formula_count:
        warnings.append(f"Detected {formula_count} formula cell(s). Scribe will not rewrite formulas automatically.")
    result = _profile_frame(frame, filename, warnings)
    result.update({"table_name": sheet.title, "format_metadata": {"sheets": [item.title for item in visible], "profiled_sheet": sheet.title, "formula_count": formula_count}, "parsing_config": parsing_config or {}})
    return result


def profile_sav_or_dta(path: Path, filename: str, suffix: str) -> dict[str, Any]:
    import pyreadstat

    if suffix == ".sav":
        frame, metadata = pyreadstat.read_sav(path, apply_value_formats=False)
    else:
        frame, metadata = pyreadstat.read_dta(path, apply_value_formats=False)
    result = _profile_frame(frame, filename)
    result.update({
        "table_name": Path(filename).stem,
        "format_metadata": {
            "column_labels": dict(zip(metadata.column_names, metadata.column_labels)),
            "value_labels": metadata.variable_value_labels,
            "original_variable_types": metadata.original_variable_types,
        },
    })
    return result


def profile_rds(path: Path, filename: str) -> dict[str, Any]:
    import pyreadr

    objects = pyreadr.read_r(path)
    if not objects:
        raise ValueError("The RDS file does not contain a readable tabular object.")
    object_name, frame = next(iter(objects.items()))
    if not hasattr(frame, "columns"):
        raise ValueError("The RDS object is not a data frame.")
    result = _profile_frame(frame, filename)
    result.update({"table_name": object_name or Path(filename).stem, "format_metadata": {"object_name": object_name or None}})
    return result


def profile_dataset(path: Path, filename: str, parsing_config: dict[str, Any] | None = None) -> dict[str, Any]:
    suffix = path.suffix.lower()
    if suffix in {".csv", ".tsv"}:
        result = profile_delimited(path, filename, parsing_config)
        result.update({"table_name": Path(filename).stem, "format_metadata": {}})
        return result
    if suffix == ".xlsx":
        return profile_xlsx(path, filename, parsing_config)
    if suffix in {".sav", ".dta"}:
        return profile_sav_or_dta(path, filename, suffix)
    if suffix == ".rds":
        return profile_rds(path, filename)
    raise ValueError(f"Unsupported dataset format: {suffix}")


def build_canonical_snapshot(item: dict[str, Any], destination: Path) -> dict[str, Any]:
    """Build a rebuildable Parquet cache with stable internal row identities."""
    import polars as pl

    frame = load_frame(Path(item["original_path"]), item["format"], item.get("profile"))
    string_columns = {
        str(column): [_plain_value(value) for value in frame[column].tolist()]
        for column in frame.columns
    }
    canonical = pl.DataFrame(string_columns, schema={name: pl.String for name in string_columns})
    canonical = canonical.with_columns(
        pl.int_range(1, canonical.height + 1, eager=True)
        .cast(pl.String)
        .map_elements(lambda ordinal: f"{item['id']}:row:{ordinal}", return_dtype=pl.String)
        .alias("_scribe_row_id")
    ).select(["_scribe_row_id", *string_columns.keys()])
    destination.parent.mkdir(parents=True, exist_ok=True)
    temporary = destination.with_suffix(destination.suffix + ".partial")
    canonical.write_parquet(temporary, compression="zstd")
    temporary.replace(destination)
    return {"path": str(destination), "rows": canonical.height, "columns": canonical.width - 1}
