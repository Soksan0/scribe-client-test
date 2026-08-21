from __future__ import annotations

import hashlib
import json
import math
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from .exporting import ordered_operations
from .formats import load_frame
from .lineage import change_row_ordinal, operation_row_ordinals, original_row_ids


SUPPORTED_OPERATION_TYPES = {
    "trim",
    "replace",
    "map_category",
    "normalize_missing",
    "parse_type",
    "parse_date",
    "derive_mapping",
    "derive_arithmetic",
    "exclude_column",
    "delete_rows",
}


@dataclass(frozen=True)
class ExpectedCell:
    row_id: str
    row_number: int
    column: str
    before: Any
    after: Any
    operation_type: str


def _is_missing(value: Any) -> bool:
    if value is None:
        return True
    try:
        return bool(math.isnan(value))
    except (TypeError, ValueError):
        return False


def canonical_value(value: Any, *, textual: bool) -> tuple[str, str]:
    """Represent values without collapsing meaningful text such as leading-zero IDs."""
    if _is_missing(value):
        return ("missing", "")
    if textual:
        return ("text", str(value))
    if isinstance(value, bool):
        return ("boolean", "true" if value else "false")
    if isinstance(value, (int, float, Decimal)):
        try:
            number = Decimal(str(value))
            return ("number", format(number.normalize(), "f"))
        except InvalidOperation:
            pass
    return ("text", str(value))


def _frame_rows(frame: Any) -> list[dict[str, Any]]:
    return [row.to_dict() for _, row in frame.iterrows()]


def expected_change_ledger(
    file_id: str,
    source_rows: list[dict[str, Any]],
    operations: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[str], list[ExpectedCell], list[str]]:
    row_ids = original_row_ids(file_id, len(source_rows))
    expected_rows = [dict(row) for row in source_rows]
    removed: set[int] = set()
    excluded_columns: list[str] = []
    ledger: dict[tuple[int, str], ExpectedCell] = {}

    for operation in ordered_operations(operations):
        operation_type = str(operation.get("type", ""))
        if operation_type not in SUPPORTED_OPERATION_TYPES:
            raise ValueError(f"Unsupported operation in validation plan: {operation_type!r}")
        if operation_type == "delete_rows":
            removed.update(operation_row_ordinals(operation))
            continue
        if operation_type == "exclude_column":
            column = str(operation.get("column", ""))
            if not expected_rows or column not in expected_rows[0]:
                raise ValueError(f"Excluded column {column!r} no longer exists")
            excluded_columns.append(column)
            for row in expected_rows:
                row.pop(column, None)
            continue
        column = str(operation.get("column", ""))
        changes = operation.get("changes", []) if operation_type in {"derive_mapping", "derive_arithmetic"} else None
        if changes is not None:
            targets = [
                (change_row_ordinal(change), change.get("before"), change.get("after"))
                for change in changes
            ]
        else:
            targets = [
                (row_number, operation.get("before"), operation.get("after"))
                for row_number in operation_row_ordinals(operation)
            ]
        for row_number, before, after in targets:
            if row_number < 1 or row_number > len(expected_rows):
                raise ValueError(f"Operation targets missing original row {row_number}")
            row = expected_rows[row_number - 1]
            if column not in row:
                raise ValueError(f"Operation targets missing column {column!r}")
            current = row[column]
            textual_current = canonical_value(current, textual=True)
            textual_before = canonical_value(before, textual=True)
            if textual_current != textual_before and canonical_value(current, textual=False) != canonical_value(before, textual=False):
                raise ValueError(
                    f"Validation plan expected {before!r} at original row {row_number}, "
                    f"column {column!r}, found {current!r}"
                )
            key = (row_number, column)
            first_before = ledger[key].before if key in ledger else current
            row[column] = after
            ledger[key] = ExpectedCell(row_ids[row_number - 1], row_number, column, first_before, after, operation_type)

    retained_rows = [row for ordinal, row in enumerate(expected_rows, start=1) if ordinal not in removed]
    retained_ids = [row_id for ordinal, row_id in enumerate(row_ids, start=1) if ordinal not in removed]
    retained_ledger = [entry for (ordinal, _), entry in ledger.items() if ordinal not in removed]
    retained_ledger.sort(key=lambda entry: (entry.row_number, entry.column))
    return retained_rows, retained_ids, retained_ledger, excluded_columns


def validate_operation_result(
    item: dict[str, Any],
    source: Path,
    output: Path,
    operations: list[dict[str, Any]],
) -> dict[str, Any]:
    profile = item.get("profile") or {}
    source_frame = load_frame(source, item["format"], profile)
    output_frame = load_frame(output, item["format"], profile)
    source_columns = [str(column) for column in source_frame.columns]
    output_columns = [str(column) for column in output_frame.columns]

    source_rows = _frame_rows(source_frame)
    output_rows = _frame_rows(output_frame)
    expected_rows, retained_ids, ledger, excluded_columns = expected_change_ledger(item["id"], source_rows, operations)
    expected_columns = [column for column in source_columns if column not in excluded_columns]
    if output_columns != expected_columns:
        raise ValueError("Cleaned output column names or order differ from the approved plan")
    if len(output_rows) != len(expected_rows):
        raise ValueError(
            f"Cleaned output has {len(output_rows)} rows; the approved plan requires {len(expected_rows)}"
        )

    textual = item["format"] in {"csv", "tsv"}
    unexpected: list[dict[str, Any]] = []
    for output_index, (expected, actual) in enumerate(zip(expected_rows, output_rows, strict=True)):
        row_id = retained_ids[output_index]
        for column in expected_columns:
            if canonical_value(actual.get(column), textual=textual) != canonical_value(expected.get(column), textual=textual):
                unexpected.append({
                    "row_id": row_id,
                    "column": column,
                    "expected": expected.get(column),
                    "actual": actual.get(column),
                })
                if len(unexpected) >= 20:
                    break
        if len(unexpected) >= 20:
            break
    if unexpected:
        first = unexpected[0]
        raise ValueError(
            "Cleaned output does not match the approved operation plan: "
            f"{first['row_id']} column {first['column']!r} expected {first['expected']!r}, "
            f"found {first['actual']!r}"
        )

    deleted_rows = len(source_rows) - len(expected_rows)
    payload = {
        "status": "passed",
        "source_rows": len(source_rows),
        "output_rows": len(output_rows),
        "source_columns": len(source_columns),
        "output_columns": len(expected_columns),
        "excluded_columns": excluded_columns,
        "expected_changed_cells": len(ledger),
        "verified_changed_cells": len(ledger),
        "unexpected_changed_cells": 0,
        "deleted_rows": deleted_rows,
        "row_order_preserved": True,
        "column_order_preserved": True,
        "ledger": [
            {
                "row_id": entry.row_id,
                "row_number": entry.row_number,
                "column": entry.column,
                "before": entry.before,
                "after": entry.after,
                "operation_type": entry.operation_type,
            }
            for entry in ledger[:100]
        ],
        "ledger_truncated": len(ledger) > 100,
    }
    canonical = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"), default=str)
    payload["validation_sha256"] = hashlib.sha256(canonical.encode("utf-8")).hexdigest()
    return payload


def validate_format_contract(item: dict[str, Any], source: Path, output: Path, operations: list[dict[str, Any]] | None = None) -> dict[str, Any]:
    """Validate metadata that is part of Scribe's supported contract for each format."""
    file_format = item["format"]
    excluded = {str(operation.get("column")) for operation in (operations or []) if operation.get("type") == "exclude_column"}
    if file_format in {"csv", "tsv"}:
        return {"status": "passed", "contract": "delimited-values", "limitations": []}
    if file_format == "xlsx":
        from openpyxl import load_workbook

        before = load_workbook(source, data_only=False, read_only=True)
        after = load_workbook(output, data_only=False, read_only=True)
        if before.sheetnames != after.sheetnames:
            raise ValueError("Workbook sheet names or order changed")
        target = (item.get("profile") or {}).get("table_name")
        for sheet_name in before.sheetnames:
            left, right = before[sheet_name], after[sheet_name]
            if sheet_name != target:
                left_values = list(left.iter_rows(values_only=True))
                right_values = list(right.iter_rows(values_only=True))
                if left_values != right_values:
                    raise ValueError(f"Untargeted worksheet {sheet_name!r} changed")
        return {"status": "passed", "contract": "xlsx-sheets-values-formulas", "limitations": ["Advanced drawing and external-link fidelity is not guaranteed by the current adapter."]}
    if file_format in {"sav", "dta"}:
        import pyreadstat

        reader = pyreadstat.read_sav if file_format == "sav" else pyreadstat.read_dta
        _, before = reader(source, apply_value_formats=False)
        _, after = reader(output, apply_value_formats=False)
        expected_names = [name for name in before.column_names if name not in excluded]
        if expected_names != after.column_names:
            raise ValueError(f"{file_format.upper()} variable names or order changed")
        expected_labels = [label for name, label in zip(before.column_names, before.column_labels, strict=True) if name not in excluded]
        if expected_labels != list(after.column_labels):
            raise ValueError(f"{file_format.upper()} variable labels changed")
        expected_value_labels = {key: value for key, value in before.variable_value_labels.items() if key not in excluded}
        if expected_value_labels != after.variable_value_labels:
            raise ValueError(f"{file_format.upper()} value labels changed")
        return {"status": "passed", "contract": f"{file_format}-labels-values", "limitations": []}
    if file_format == "rds":
        before = load_frame(source, file_format, item.get("profile"))
        after = load_frame(output, file_format, item.get("profile"))
        expected_columns = [str(value) for value in before.columns if str(value) not in excluded]
        if expected_columns != [str(value) for value in after.columns]:
            raise ValueError("RDS columns or order changed")
        return {"status": "passed", "contract": "rds-data-frame", "limitations": ["Only data-frame values, columns, order, and representable pandas dtypes are verified."]}
    raise ValueError(f"No preservation contract exists for {file_format!r}")
