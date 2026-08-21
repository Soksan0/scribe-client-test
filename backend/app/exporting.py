from __future__ import annotations

import csv
import hashlib
import io
import json
import os
import shutil
from pathlib import Path
from typing import Any

from .lineage import change_row_ordinal, operation_row_ordinals
from .profiling import decode_bytes, detect_delimiter


OPERATION_PRIORITY = {
    "trim": 10,
    "normalize_missing": 20,
    "map_category": 30,
    "derive_mapping": 35,
    "parse_date": 40,
    "parse_type": 40,
    "derive_arithmetic": 45,
    "replace": 50,
    "exclude_column": 90,
    "delete_rows": 100,
}


def ordered_operations(operations: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Apply deterministic normalizations before corrections that depend on them."""
    indexed = list(enumerate(coalesce_operations(operations)))
    def priority(item: tuple[int, dict[str, Any]]) -> tuple[int, int]:
        index, operation = item
        base = OPERATION_PRIORITY.get(operation.get("type", ""), 90)
        if operation.get("type") in {"derive_mapping", "derive_arithmetic"}:
            base += max(0, int(operation.get("phase", 1)) - 1) * 20
        return base, index
    indexed.sort(key=priority)
    return [operation for _, operation in indexed]


def coalesce_operations(operations: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Merge repeated logical corrections so each source cell is transformed once."""
    grouped: dict[str, dict[str, Any]] = {}
    order: list[str] = []
    for operation in operations:
        base = {key: value for key, value in operation.items() if key not in {"rows", "row_ids"}}
        key = json.dumps(base, ensure_ascii=False, sort_keys=True, default=str)
        if key not in grouped:
            grouped[key] = {**operation, "rows": [], "row_ids": []}
            order.append(key)
        grouped[key]["rows"] = sorted(set(grouped[key].get("rows", [])) | {int(row) for row in operation.get("rows", [])})
        grouped[key]["row_ids"] = sorted(set(grouped[key].get("row_ids", [])) | {str(row_id) for row_id in operation.get("row_ids", [])})
    return [grouped[key] for key in order]


def replace_exact_value(row: dict[str, Any], column: str, expected: Any, replacement: Any, row_number: int) -> None:
    current = row.get(column)
    if current == replacement:
        return
    if current != expected:
        raise ValueError(
            f"Conflicting accepted corrections at row {row_number}, column {column}: "
            f"expected {expected!r}, found {current!r}"
        )
    row[column] = replacement


def values_equivalent(current: Any, expected: Any) -> bool:
    if current == expected:
        return True
    if current is None and expected in {None, ""}:
        return True
    try:
        return float(str(current).replace(",", "")) == float(str(expected).replace(",", ""))
    except (TypeError, ValueError):
        return str(current) == str(expected)


def apply_derived_change(row: dict[str, Any], operation: dict[str, Any], change: dict[str, Any]) -> None:
    row_number = change_row_ordinal(change)
    column = operation["column"]
    current = row.get(column)
    if values_equivalent(current, change["after"]):
        return
    if not values_equivalent(current, change["before"]):
        raise ValueError(f"Derived correction at row {row_number}, column {column} expected {change['before']!r}, found {current!r}")
    for source_column, expected in change.get("inputs", {}).items():
        if not values_equivalent(row.get(source_column), expected):
            raise ValueError(f"Derived correction at row {row_number} requires {source_column}={expected!r}, found {row.get(source_column)!r}")
    row[column] = change["after"]


def r_literal(value: Any) -> str:
    if value is None:
        return "NA"
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, str):
        return json.dumps(value)
    return str(value)


def append_r_operations(lines: list[str], operations: list[dict[str, Any]], csv_reader: bool) -> None:
    parse_columns: dict[str, str] = {}
    delete_rows: set[int] = set()
    step = 0
    for operation in ordered_operations(operations):
        if operation.get("type") == "delete_rows":
            delete_rows.update(operation_row_ordinals(operation))
            continue
        if operation.get("type") == "exclude_column":
            step += 1
            column = json.dumps(operation["column"])
            lines.extend(["", f"# Step {step}: explicitly approved exclusion of column {operation['column']}", f"stopifnot({column} %in% names(data))", f"data <- data[, setdiff(names(data), {column}), drop = FALSE]"])
            continue
        if operation.get("type") in {"derive_mapping", "derive_arithmetic"}:
            step += 1
            column = json.dumps(operation["column"])
            changes = operation.get("changes", [])
            lines.extend(["", f"# Step {step}: accepted {operation['type']} using {operation.get('formula', 'verified row evidence')}", "# Preconditions stop the script if source values no longer match the reviewed evidence."])
            for change in changes:
                row_number = change_row_ordinal(change)
                before = json.dumps(str(change.get("before", "")))
                lines.append(f"stopifnot(as.character(data[[{column}]][{row_number}]) == {before})")
                for source_column, expected in change.get("inputs", {}).items():
                    source = json.dumps(source_column)
                    lines.append(f"stopifnot(as.character(data[[{source}]][{row_number}]) == {json.dumps(str(expected))})")
                lines.append(f"data[[{column}]][{row_number}] <- {r_literal(change.get('after'))}")
            continue
        if operation.get("type") not in {"trim", "replace", "map_category", "normalize_missing", "parse_type", "parse_date"}:
            continue
        step += 1
        column = json.dumps(operation["column"])
        rows = ", ".join(str(row) for row in operation_row_ordinals(operation)) or "integer(0)"
        before = json.dumps(str(operation.get("before", "")))
        after = json.dumps(str(operation.get("after", ""))) if operation.get("type") == "parse_type" else r_literal(operation.get("after"))
        lines.extend([
            "",
            f"# Step {step}: accepted {operation['type']} correction on exact source rows",
            f"target_rows <- c({rows})",
            f"matching_rows <- target_rows[as.character(data[[{column}]][target_rows]) == {before}]",
            "stopifnot(length(matching_rows) == length(target_rows))",
            f"data[[{column}]][matching_rows] <- {after}",
        ])
        if operation.get("type") == "parse_type":
            parse_columns[operation["column"]] = operation.get("expected", "number")
    for column_name, expected in parse_columns.items():
        column = json.dumps(column_name)
        converter = ("readr::parse_integer" if expected == "integer" else "readr::parse_number") if csv_reader else ("as.integer" if expected == "integer" else "as.numeric")
        if csv_reader:
            lines.extend([
                "",
                f"# Convert {column_name} after all reviewed token replacements",
                f"parsed_values <- {converter}(as.character(data[[{column}]]), na = character())",
                "stopifnot(nrow(readr::problems(parsed_values)) == 0)",
                f"data[[{column}]] <- parsed_values",
            ])
        else:
            lines.extend(["", f"# Convert {column_name} after all reviewed token replacements", f"data[[{column}]] <- {converter}(as.character(data[[{column}]]))"])
    if delete_rows:
        rows = ", ".join(str(row) for row in sorted(delete_rows))
        lines.extend(["", "# Remove explicitly approved duplicate rows after cell corrections", f"data <- data[-c({rows}), , drop = FALSE]"])


def apply_operations(source: Path, destination: Path, operations: list[dict[str, Any]], parsing_context: dict[str, Any] | None = None) -> dict[str, Any]:
    content = source.read_bytes()
    context = parsing_context or {}
    requested_encoding = context.get("encoding")
    if requested_encoding:
        text, encoding = content.decode(str(requested_encoding)), str(requested_encoding)
    else:
        text, encoding, _ = decode_bytes(content)
    delimiter = str(context.get("delimiter") or detect_delimiter(text, source.name))
    header_row = int(context.get("header_row") or 1)
    if header_row > 1:
        raw_rows = list(csv.reader(io.StringIO(text), delimiter=delimiter, strict=True))
        if header_row > len(raw_rows):
            raise ValueError(f"Header row {header_row} is outside the source file")
        staged = io.StringIO()
        csv.writer(staged, delimiter=delimiter, lineterminator="\n").writerows(raw_rows[header_row - 1 :])
        text = staged.getvalue()
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    headers = reader.fieldnames or []
    rows = list(reader)
    delete_rows: set[int] = set()
    for operation in ordered_operations(operations):
        if operation.get("type") == "delete_rows":
            delete_rows.update(operation_row_ordinals(operation))
            continue
        if operation.get("type") == "exclude_column":
            column = operation["column"]
            if column not in headers:
                raise ValueError(f"Excluded column {column!r} no longer exists")
            headers.remove(column)
            for row in rows:
                row.pop(column, None)
            continue
        if operation.get("type") in {"derive_mapping", "derive_arithmetic"}:
            column = operation["column"]
            for change in operation.get("changes", []):
                row_number = change_row_ordinal(change)
                row_index = row_number - 1
                if row_index < 0 or row_index >= len(rows):
                    raise ValueError(f"Derived correction refers to missing row {row_number}, column {column}")
                apply_derived_change(rows[row_index], operation, change)
            continue
        if operation.get("type") not in {"trim", "replace", "map_category", "normalize_missing", "parse_type", "parse_date"}:
            continue
        column, expected, replacement = operation["column"], operation["before"], operation["after"]
        for row_number in operation_row_ordinals(operation):
            row_index = row_number - 1
            if row_index < 0 or row_index >= len(rows):
                raise ValueError(f"Correction refers to missing row {row_number}, column {column}")
            replace_exact_value(rows[row_index], column, expected, replacement, row_number)
    rows = [row for index, row in enumerate(rows, start=1) if index not in delete_rows]
    destination.parent.mkdir(parents=True, exist_ok=True)
    temporary = destination.with_name(f"{destination.stem}.partial{destination.suffix}")
    with temporary.open("w", encoding=encoding, newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers, delimiter=delimiter, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)
        handle.flush()
        os.fsync(handle.fileno())
    temporary.replace(destination)
    output = destination.read_bytes()
    return {"path": str(destination), "sha256": hashlib.sha256(output).hexdigest(), "size_bytes": len(output)}


def generate_r_script(
    source_name: str,
    output_name: str,
    operations: list[dict[str, Any]],
    delimiter: str,
    parsing_context: dict[str, Any] | None = None,
) -> str:
    context = parsing_context or {}
    encoding = json.dumps(context.get("encoding") or "UTF-8")
    delimiter_literal = json.dumps(delimiter)
    source_literal = json.dumps(f"originals/{source_name}")
    output_literal = json.dumps(f"cleaned/{output_name}")
    expected_columns = [str(item.get("name")) for item in context.get("columns", []) if item.get("name") is not None]
    header_row = int((context.get("parsing_config") or {}).get("header_row") or 1)
    expected_columns_r = ", ".join(json.dumps(value) for value in expected_columns)
    lines = [
        "# Generated by Scribe. Review before running.",
        "# Every transformation corresponds to an accepted audit decision.",
        "library(readr)",
        "library(dplyr)",
        "",
        "# Import every source column as text so identifiers and exact preconditions are preserved.",
        f"data <- readr::read_delim({source_literal}, delim = {delimiter_literal}, skip = {header_row - 1}, col_types = readr::cols(.default = readr::col_character()), locale = readr::locale(encoding = {encoding}), na = character(), trim_ws = FALSE, name_repair = \"minimal\", skip_empty_rows = FALSE, show_col_types = FALSE, progress = FALSE)",
        "stopifnot(nrow(readr::problems(data)) == 0)",
    ]
    if expected_columns:
        lines.append(f"stopifnot(identical(names(data), c({expected_columns_r})))")
    append_r_operations(lines, operations, csv_reader=True)
    lines.extend(["", f"readr::write_delim(data, {output_literal}, delim = {delimiter_literal}, na = \"\")", ""])
    return "\n".join(lines)


def apply_xlsx_operations(source: Path, destination: Path, operations: list[dict[str, Any]], sheet_name: str, header_row: int = 1) -> dict[str, Any]:
    from openpyxl import load_workbook

    destination.parent.mkdir(parents=True, exist_ok=True)
    temporary = destination.with_name(f"{destination.stem}.partial{destination.suffix}")
    shutil.copy2(source, temporary)
    workbook = load_workbook(temporary)
    sheet = workbook[sheet_name]
    headers = {str(cell.value): cell.column for cell in sheet[header_row] if cell.value is not None}
    delete_rows: set[int] = set()
    for operation in ordered_operations(operations):
        if operation.get("type") == "delete_rows":
            delete_rows.update(operation_row_ordinals(operation))
            continue
        if operation.get("type") == "exclude_column":
            column = operation["column"]
            if column not in headers:
                raise ValueError(f"Excluded column {column!r} no longer exists in worksheet {sheet_name!r}")
            sheet.delete_cols(headers[column])
            headers = {str(cell.value): cell.column for cell in sheet[header_row] if cell.value is not None}
            continue
        if operation.get("type") in {"derive_mapping", "derive_arithmetic"}:
            column = operation["column"]
            if column not in headers:
                raise ValueError(f"Column {column!r} no longer exists in worksheet {sheet_name!r}")
            for change in operation.get("changes", []):
                row_number = change_row_ordinal(change)
                row_values = {name: sheet.cell(row=row_number + header_row, column=position).value for name, position in headers.items()}
                apply_derived_change(row_values, operation, change)
                sheet.cell(row=row_number + header_row, column=headers[column]).value = row_values[column]
            continue
        if operation.get("type") not in {"trim", "replace", "map_category", "normalize_missing", "parse_type", "parse_date"}:
            continue
        column = operation["column"]
        if column not in headers:
            raise ValueError(f"Column {column!r} no longer exists in worksheet {sheet_name!r}")
        for row_number in operation_row_ordinals(operation):
            cell = sheet.cell(row=row_number + header_row, column=headers[column])
            if cell.data_type == "f":
                raise ValueError(f"Scribe will not replace a formula at row {row_number}, column {column}")
            current = cell.value
            if current == operation["after"]:
                continue
            if current != operation["before"]:
                raise ValueError(
                    f"Conflicting accepted corrections at row {row_number}, column {column}: "
                    f"expected {operation['before']!r}, found {current!r}"
                )
            cell.value = operation["after"]
    for row_number in sorted(delete_rows, reverse=True):
        sheet.delete_rows(row_number + header_row)
    workbook.save(temporary)
    temporary.replace(destination)
    output = destination.read_bytes()
    return {"path": str(destination), "sha256": hashlib.sha256(output).hexdigest(), "size_bytes": len(output)}


def apply_statistical_operations(source: Path, destination: Path, operations: list[dict[str, Any]], file_format: str) -> dict[str, Any]:
    import pyreadr
    import pyreadstat

    metadata = None
    if file_format == "sav":
        frame, metadata = pyreadstat.read_sav(source, apply_value_formats=False)
    elif file_format == "dta":
        frame, metadata = pyreadstat.read_dta(source, apply_value_formats=False)
    else:
        objects = pyreadr.read_r(source)
        if not objects:
            raise ValueError("The RDS file does not contain a readable data frame")
        frame = next(iter(objects.values()))
    delete_rows: set[int] = set()
    for operation in ordered_operations(operations):
        if operation.get("type") == "delete_rows":
            delete_rows.update(operation_row_ordinals(operation))
            continue
        if operation.get("type") == "exclude_column":
            column = operation["column"]
            if column not in frame.columns:
                raise ValueError(f"Excluded column {column!r} no longer exists")
            frame = frame.drop(columns=[column])
            continue
        if operation.get("type") in {"derive_mapping", "derive_arithmetic"}:
            column = operation["column"]
            for change in operation.get("changes", []):
                row_number = change_row_ordinal(change)
                row_index = row_number - 1
                if row_index < 0 or row_index >= len(frame):
                    raise ValueError(f"Derived correction refers to missing row {row_number}, column {column}")
                row_values = frame.iloc[row_index].to_dict()
                apply_derived_change(row_values, operation, change)
                frame.at[frame.index[row_index], column] = row_values[column]
            continue
        if operation.get("type") not in {"trim", "replace", "map_category", "normalize_missing", "parse_type", "parse_date"}:
            continue
        column = operation["column"]
        for row_number in operation_row_ordinals(operation):
            row_index = row_number - 1
            if row_index < 0 or row_index >= len(frame):
                raise ValueError(f"Correction refers to missing row {row_number}, column {column}")
            current = frame.iloc[row_index][column]
            if current == operation["after"]:
                continue
            if current != operation["before"]:
                raise ValueError(
                    f"Conflicting accepted corrections at row {row_number}, column {column}: "
                    f"expected {operation['before']!r}, found {current!r}"
                )
            frame.at[frame.index[row_index], column] = operation["after"]
    if delete_rows:
        frame = frame.drop(frame.index[[row - 1 for row in sorted(delete_rows)]]).reset_index(drop=True)
    destination.parent.mkdir(parents=True, exist_ok=True)
    temporary = destination.with_name(f"{destination.stem}.partial{destination.suffix}")
    if file_format == "sav":
        labels = {name: label for name, label in zip(metadata.column_names, metadata.column_labels, strict=True) if name in frame.columns}
        value_labels = {name: labels for name, labels in metadata.variable_value_labels.items() if name in frame.columns}
        pyreadstat.write_sav(frame, temporary, column_labels=labels, variable_value_labels=value_labels)
    elif file_format == "dta":
        labels = {name: label for name, label in zip(metadata.column_names, metadata.column_labels, strict=True) if name in frame.columns}
        value_labels = {name: labels for name, labels in metadata.variable_value_labels.items() if name in frame.columns}
        pyreadstat.write_dta(frame, temporary, column_labels=labels, variable_value_labels=value_labels)
    else:
        pyreadr.write_rds(temporary, frame)
    temporary.replace(destination)
    output = destination.read_bytes()
    return {"path": str(destination), "sha256": hashlib.sha256(output).hexdigest(), "size_bytes": len(output)}


def generate_format_r_script(source_name: str, output_name: str, operations: list[dict[str, Any]], file_format: str, delimiter: str | None = None, sheet_name: str | None = None, parsing_context: dict[str, Any] | None = None) -> str:
    if file_format in {"csv", "tsv"}:
        return generate_r_script(source_name, output_name, operations, delimiter or ("\t" if file_format == "tsv" else ","), parsing_context)
    if file_format == "xlsx":
        libraries = ["library(openxlsx)", "library(dplyr)"]
        sheet = json.dumps(sheet_name or 1)
        header_row = int(((parsing_context or {}).get("parsing_config") or {}).get("header_row") or 1)
        read_line = f'workbook <- openxlsx::loadWorkbook("originals/{source_name}")\ndata <- openxlsx::read.xlsx(workbook, sheet = {sheet}, startRow = {header_row}, check.names = FALSE, skipEmptyRows = FALSE, skipEmptyCols = FALSE)\noriginal_data_rows <- nrow(data)\noriginal_data_columns <- ncol(data)'
        write_line = f'openxlsx::deleteData(workbook, sheet = {sheet}, cols = seq_len(original_data_columns), rows = {header_row}:({header_row} + original_data_rows))\nopenxlsx::writeData(workbook, sheet = {sheet}, startRow = {header_row}, x = data)\nopenxlsx::saveWorkbook(workbook, "cleaned/{output_name}", overwrite = TRUE)'
    elif file_format in {"sav", "dta"}:
        libraries = ["library(haven)", "library(dplyr)"]
        read_line = f'data <- haven::read_{file_format}("originals/{source_name}")'
        write_line = f'haven::write_{file_format}(data, "cleaned/{output_name}")'
    else:
        libraries = ["library(dplyr)"]
        read_line = f'data <- readRDS("originals/{source_name}")'
        write_line = f'saveRDS(data, "cleaned/{output_name}")'
    lines = ["# Generated by Scribe. Review before running.", "# Every transformation corresponds to an accepted audit decision.", *libraries, "", read_line]
    append_r_operations(lines, operations, csv_reader=False)
    lines.extend(["", write_line, ""])
    return "\n".join(lines)
