from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from backend.app.exporting import apply_operations, apply_xlsx_operations, coalesce_operations, generate_format_r_script, generate_r_script
from backend.app.profiling import profile_delimited
from backend.app.formats import profile_dataset
from backend.app.validation import validate_format_contract, validate_operation_result
from backend.app.survey_quality import detect_survey_quality


class ProfilingTests(unittest.TestCase):
    def test_survey_quality_checks_require_configuration_and_never_propose_edits(self) -> None:
        import pandas as pd

        frame = pd.DataFrame({
            "participant_id": ["P1", "P1", "P2"],
            "q1": [1, 1, 5], "q2": [1, 1, 4], "q3": [1, 1, 3],
            "attention": [2, 1, 2], "duration": [10, 12, 100],
        })
        self.assertEqual(detect_survey_quality(frame, {}), [])
        findings = detect_survey_quality(frame, {
            "participant_keys": ["participant_id"], "allowed_repeats": False,
            "item_groups": [{"name": "Scale", "columns": ["q1", "q2", "q3"], "minimum": 1, "maximum": 5}],
            "attention_checks": [{"column": "attention", "expected_values": [2]}],
            "timestamp_columns": {"duration": "duration"},
        })
        categories = {item["category"] for item in findings}
        self.assertTrue({"survey_duplicate_submission", "survey_straightlining", "survey_attention_check"}.issubset(categories))
        self.assertTrue(all("operation" not in item for item in findings))

    def test_derives_only_proven_transaction_values_and_excludes_ambiguous_reverse_mapping(self) -> None:
        import csv

        rows = []
        for index in range(10):
            rows.append([f"C{index}", "Coffee", 2, 2, 4])
            rows.append([f"T{index}", "Tea", 2, 1.5, 3])
        for index in range(5):
            rows.append([f"S{index}", "Sandwich", 2, 4, 8])
            rows.append([f"M{index}", "Smoothie", 2, 4, 8])
        rows.extend([
            ["Q", "Smoothie", "ERROR", 4, 20],
            ["P", "Coffee", 3, "UNKNOWN", 6],
            ["L", "Tea", 2, 1.5, "ERROR"],
            ["I", "UNKNOWN", 2, 2, 4],
            ["A", "UNKNOWN", 2, 4, 8],
        ])
        with tempfile.TemporaryDirectory() as folder:
            source = Path(folder) / "sales.csv"
            reviewed = Path(folder) / "reviewed.csv"
            with source.open("w", encoding="utf-8", newline="") as handle:
                writer = csv.writer(handle)
                writer.writerow(["Transaction ID", "Item", "Quantity", "Price Per Unit", "Total Spent"])
                writer.writerows(rows)
            profile = profile_delimited(source, source.name)
            operations = [item["operation"] for item in profile["findings"] if item["operation"] and item["confidence"] == "high"]
            apply_operations(source, reviewed, operations)
            with reviewed.open(encoding="utf-8-sig", newline="") as handle:
                output = {row["Transaction ID"]: row for row in csv.DictReader(handle)}
        self.assertEqual(output["Q"]["Quantity"], "5")
        self.assertEqual(output["P"]["Price Per Unit"], "2")
        self.assertEqual(output["L"]["Total Spent"], "3")
        self.assertEqual(output["I"]["Item"], "Coffee")
        self.assertEqual(output["A"]["Item"], "")
        self.assertIn("arithmetic_inference", {item["category"] for item in profile["findings"]})
        self.assertIn("codebook_inference", {item["category"] for item in profile["findings"]})

    def test_derived_operation_checks_every_source_precondition(self) -> None:
        operation = {
            "type": "derive_arithmetic", "column": "Quantity", "formula": "Total / Price",
            "source_columns": ["Total", "Price"], "rows": [1],
            "changes": [{"row": 1, "before": "ERROR", "after": 5, "inputs": {"Total": "20", "Price": "4"}}],
        }
        with tempfile.TemporaryDirectory() as folder:
            source, reviewed = Path(folder) / "source.csv", Path(folder) / "reviewed.csv"
            source.write_text("Quantity,Price,Total\nERROR,5,20\n", encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "requires Price"):
                apply_operations(source, reviewed, [operation])

    def test_one_arithmetic_contradiction_blocks_all_automatic_derivation(self) -> None:
        rows = [f"T{index},Coffee,2,2,4" for index in range(10)]
        rows.extend(["DISCOUNT,Coffee,2,2,3", "MISSING,Coffee,ERROR,2,4"])
        content = "Transaction ID,Item,Quantity,Price Per Unit,Total Spent\n" + "\n".join(rows) + "\n"
        with tempfile.TemporaryDirectory() as folder:
            source = Path(folder) / "discounts.csv"
            source.write_text(content, encoding="utf-8")
            profile = profile_delimited(source, source.name)
        categories = {item["category"] for item in profile["findings"]}
        self.assertIn("arithmetic_inconsistency", categories)
        self.assertNotIn("arithmetic_inference", categories)

    def test_generated_r_script_records_formula_and_preconditions_for_derived_values(self) -> None:
        operation = {
            "type": "derive_arithmetic", "column": "Quantity", "formula": "Total / Price",
            "source_columns": ["Total", "Price"], "rows": [1],
            "changes": [{"row": 1, "before": "ERROR", "after": 5, "inputs": {"Total": "20", "Price": "4"}}],
        }
        script = generate_r_script("source.csv", "reviewed.csv", [operation], ",")
        self.assertIn("accepted derive_arithmetic using Total / Price", script)
        self.assertIn("stopifnot", script)
        self.assertIn('data[["Quantity"]][1] <- 5', script)

    def test_real_cafe_fixture_recovers_all_provable_values_without_guessing(self) -> None:
        import csv
        import hashlib

        source = Path("/Users/soksanhay/Downloads/dirty_cafe_sales.csv")
        if not source.exists():
            self.skipTest("The user-provided cafe acceptance fixture is not available")
        original_hash = hashlib.sha256(source.read_bytes()).hexdigest()
        profile = profile_delimited(source, source.name)
        operations = coalesce_operations([item["operation"] for item in profile["findings"] if item["operation"] and item["confidence"] == "high"])
        derived = [item for item in profile["findings"] if item["category"] in {"arithmetic_inference", "codebook_inference"}]
        self.assertEqual((profile["row_count"], profile["column_count"]), (10_000, 8))
        self.assertEqual(profile["candidate_id_columns"], ["Transaction ID"])
        self.assertEqual(sum(item["affected_count"] for item in derived if item["category"] == "arithmetic_inference"), 983)
        self.assertEqual(sum(item["affected_count"] for item in derived if item["category"] == "codebook_inference"), 968)
        self.assertEqual(sum(len(item.get("rows", [])) for item in operations), 3_887)
        with tempfile.TemporaryDirectory() as folder:
            reviewed = Path(folder) / source.name
            apply_operations(source, reviewed, operations)
            with reviewed.open(encoding="utf-8-sig", newline="") as handle:
                rows = list(csv.DictReader(handle))
        example = next(row for row in rows if row["Transaction ID"] == "TXN_3522028")
        self.assertEqual(example["Quantity"], "5")
        self.assertEqual(len(rows), 10_000)
        self.assertFalse(any(value in {"ERROR", "UNKNOWN"} for row in rows for value in row.values()))
        self.assertFalse(any(
            abs(float(row["Quantity"]) * float(row["Price Per Unit"]) - float(row["Total Spent"])) > 1e-9
            for row in rows if row["Quantity"] and row["Price Per Unit"] and row["Total Spent"]
        ))
        self.assertEqual(hashlib.sha256(source.read_bytes()).hexdigest(), original_hash)

    def test_detects_whitespace_duplicate_ids_and_rows(self) -> None:
        content = "participant_id,occupation,age\nP001,Teacher  ,30\nP002,Nurse,29\nP002,Nurse,29\n"
        with tempfile.TemporaryDirectory() as folder:
            path = Path(folder) / "study.csv"
            path.write_text(content, encoding="utf-8")
            result = profile_delimited(path, path.name)
        self.assertEqual(result["row_count"], 3)
        self.assertEqual(result["column_count"], 3)
        self.assertEqual({finding["category"] for finding in result["findings"]}, {"whitespace", "duplicate_id", "duplicate_row"})

    def test_detects_missing_values_and_category_case_without_guessing(self) -> None:
        content = "participant_id,status,age\nP001,Active,20\nP002,active,\nP003,Active,22\n"
        with tempfile.TemporaryDirectory() as folder:
            path = Path(folder) / "study.csv"
            path.write_text(content, encoding="utf-8")
            result = profile_delimited(path, path.name)
        by_category = {finding["category"]: finding for finding in result["findings"]}
        self.assertIn("missing_value", by_category)
        self.assertIn("inconsistent_category", by_category)
        self.assertIsNone(by_category["missing_value"]["proposed"])
        self.assertEqual(by_category["inconsistent_category"]["proposed"], "Active")
        self.assertEqual(by_category["inconsistent_category"]["operation"]["type"], "map_category")

    def test_detects_nan_number_words_impossible_ages_and_duplicate_columns(self) -> None:
        content = "participant_id,age,copy_a,copy_b\nP001,30,X,X\nP002,forty,Y,Y\nP003,nan,Z,Z\nP004,121,Q,Q\n"
        with tempfile.TemporaryDirectory() as folder:
            path = Path(folder) / "ages.csv"
            path.write_text(content, encoding="utf-8")
            result = profile_delimited(path, path.name)
        categories = {finding["category"] for finding in result["findings"]}
        self.assertTrue({"missing_value", "invalid_type", "impossible_value", "duplicate_column"}.issubset(categories))
        conversion = next(finding for finding in result["findings"] if finding["category"] == "invalid_type")
        self.assertEqual(conversion["before"], "forty")
        self.assertEqual(conversion["proposed"], 40)
        self.assertEqual(conversion["operation"]["expected"], "integer")

    def test_contextual_na_is_not_missing_for_state_codes(self) -> None:
        content = "participant_id,state,age\nP001,NA,30\nP002,CA,nan\n"
        with tempfile.TemporaryDirectory() as folder:
            path = Path(folder) / "states.csv"
            path.write_text(content, encoding="utf-8")
            result = profile_delimited(path, path.name)
        state_profile = next(column for column in result["columns"] if column["name"] == "state")
        age_profile = next(column for column in result["columns"] if column["name"] == "age")
        self.assertEqual(state_profile["missing_count"], 0)
        self.assertEqual(age_profile["missing_count"], 1)

    def test_ambiguous_dates_are_review_only(self) -> None:
        content = "participant_id,visit_date\nP001,03/04/2020\nP002,2020-05-06\nP003,13/04/2020\n"
        with tempfile.TemporaryDirectory() as folder:
            path = Path(folder) / "dates.csv"
            path.write_text(content, encoding="utf-8")
            result = profile_delimited(path, path.name)
        ambiguous = next(finding for finding in result["findings"] if finding["category"] == "ambiguous_date")
        self.assertEqual(ambiguous["before"], "03/04/2020")
        self.assertIsNone(ambiguous["operation"])
        standardized = [finding for finding in result["findings"] if finding["category"] == "date_standardization"]
        self.assertTrue(any(item["before"] == "13/04/2020" and item["proposed"] == "2020-04-13" for item in standardized))

    def test_detects_patterns_boolean_variants_and_invalid_scales(self) -> None:
        content = "participant_id,email,phone,blood_pressure,consent,satisfaction_score\nP001,valid@example.com,5551234567,120/80,Yes,5\nP002,not-an-email,12,80/120,Y,0\nP003,also@valid.org,555-222-3333,118/75,TRUE,4\nP004,third@valid.org,5553334444,121/78,1,6\nP005,fourth@valid.org,5553334445,119/79,no,3\nP006,fifth@valid.org,5553334446,117/76,N,2\nP007,sixth@valid.org,5553334447,122/82,false,1\n"
        with tempfile.TemporaryDirectory() as folder:
            path = Path(folder) / "patterns.csv"
            path.write_text(content, encoding="utf-8")
            result = profile_delimited(path, path.name)
        categories = {finding["category"] for finding in result["findings"]}
        self.assertIn("invalid_pattern", categories)
        self.assertIn("boolean_standardization", categories)
        self.assertIn("invalid_scale", categories)

    def test_identifier_phone_and_code_columns_stay_text(self) -> None:
        content = "participant_id,phone,zip_code,age\n00123,555-222-3333,02108,thirty\n00124,5553334444,10001,40\n"
        with tempfile.TemporaryDirectory() as folder:
            path = Path(folder) / "semantic_text.csv"
            path.write_text(content, encoding="utf-8")
            result = profile_delimited(path, path.name)
        by_column = {column["name"]: column for column in result["columns"]}
        self.assertEqual(by_column["participant_id"]["inferred_type"], "text")
        self.assertEqual(by_column["phone"]["inferred_type"], "text")
        self.assertEqual(by_column["zip_code"]["inferred_type"], "text")
        conversions = [finding for finding in result["findings"] if finding["category"] == "invalid_type"]
        self.assertEqual(len(conversions), 1)
        self.assertEqual(conversions[0]["column_name"], "age")
        self.assertEqual(conversions[0]["proposed"], 30)

    def test_duplicate_headers_stop_ambiguous_cell_level_cleaning(self) -> None:
        content = "ID,ID\nP001,A\nP002,A\n"
        with tempfile.TemporaryDirectory() as folder:
            path = Path(folder) / "duplicate_headers.csv"
            path.write_text(content, encoding="utf-8")
            result = profile_delimited(path, path.name)
        self.assertEqual({finding["category"] for finding in result["findings"]}, {"duplicate_column_name"})
        self.assertEqual(result["row_count"], 2)

    def test_profiles_ten_thousand_rows_and_two_hundred_columns_under_reference_time(self) -> None:
        import csv
        import time
        columns = ["participant_id", *[f"measure_{index}" for index in range(200)]]
        with tempfile.TemporaryDirectory() as folder:
            path = Path(folder) / "large.csv"
            with path.open("w", encoding="utf-8", newline="") as handle:
                writer = csv.writer(handle)
                writer.writerow(columns)
                for row in range(10_001):
                    writer.writerow([f"P{row:05d}", *[(row + column) % 17 for column in range(200)]])
            started = time.monotonic()
            result = profile_delimited(path, path.name)
            elapsed = time.monotonic() - started
        self.assertEqual(result["row_count"], 10_001)
        self.assertEqual(result["column_count"], 201)
        self.assertLess(elapsed, 60)

    def test_real_healthcare_fixture_produces_analyzable_reviewed_copy(self) -> None:
        import csv
        import hashlib

        source = Path(__file__).resolve().parents[2] / "healthcare_messy_data.csv"
        self.assertTrue(source.exists(), "The attached healthcare acceptance fixture is required")
        original_hash = hashlib.sha256(source.read_bytes()).hexdigest()
        profile = profile_delimited(source, source.name)
        self.assertEqual((profile["row_count"], profile["column_count"]), (1000, 10))
        age_conversion = next(item for item in profile["findings"] if item["category"] == "invalid_type" and item["column_name"] == "Age")
        self.assertEqual((age_conversion["before"], age_conversion["proposed"], age_conversion["affected_count"]), ("forty", 40, 176))
        self.assertEqual(len([item for item in profile["findings"] if item["category"] == "date_standardization"]), 5)
        self.assertEqual(len([item for item in profile["findings"] if item["category"] == "potential_duplicate"]), 2)
        by_column = {item["name"]: item for item in profile["columns"]}
        self.assertEqual(by_column["Condition"]["missing_count"], 0)
        self.assertEqual(by_column["Medication"]["missing_count"], 0)

        accepted = [item["operation"] for item in profile["findings"] if item["operation"] and item["confidence"] == "high"]
        coalesced = coalesce_operations(accepted)
        touched_cells = {
            (row_number, operation["column"])
            for operation in coalesced
            if operation.get("column")
            for row_number in operation.get("rows", [])
        }
        self.assertEqual(len(coalesced), 22)
        self.assertEqual(len(touched_cells), 3_311)
        self.assertEqual(
            sum(1 for operation in coalesced if operation["type"] == "parse_type" and operation.get("column") == "Age"),
            1,
        )
        self.assertIn("identity_not_verifiable", {item["category"] for item in profile["findings"]})
        self.assertFalse(any(item["category"] == "duplicate_id" and item["column_name"] == "Patient Name" for item in profile["findings"]))
        with tempfile.TemporaryDirectory() as folder:
            reviewed = Path(folder) / "healthcare_messy_data.csv"
            apply_operations(source, reviewed, coalesced)
            with reviewed.open(newline="", encoding="utf-8-sig") as handle:
                rows = list(csv.DictReader(handle))
        self.assertEqual(len(rows), 1000)
        self.assertFalse(any(value.strip().casefold() in {"nan", "forty"} for row in rows for value in row.values()))
        self.assertTrue(all(row["Age"] == "" or row["Age"].isdigit() for row in rows))
        self.assertTrue(all(row["Visit Date"][:4].isdigit() and row["Visit Date"].count("-") == 2 for row in rows))
        self.assertTrue(all(row["Patient Name"] == row["Patient Name"].strip() for row in rows))
        self.assertEqual(rows[0]["Patient Name"], "david lee")
        self.assertIn("None", {row["Condition"] for row in rows})
        self.assertIn("NONE", {row["Medication"] for row in rows})
        self.assertEqual(hashlib.sha256(source.read_bytes()).hexdigest(), original_hash)

    def test_export_applies_only_explicit_operations(self) -> None:
        with tempfile.TemporaryDirectory() as folder:
            source, destination = Path(folder) / "source.csv", Path(folder) / "reviewed.csv"
            source.write_text("id,value\n1,Teacher  \n2,Nurse\n", encoding="utf-8")
            operation = {"type": "trim", "column": "value", "before": "Teacher  ", "after": "Teacher", "rows": [1]}
            result = apply_operations(source, destination, [operation])
            self.assertIn("1,Teacher\n", destination.read_text(encoding="utf-8"))
            self.assertTrue(result["sha256"])
            self.assertEqual(source.read_text(encoding="utf-8"), "id,value\n1,Teacher  \n2,Nurse\n")

    def test_csv_export_preserves_quoted_multiline_cells(self) -> None:
        import csv
        with tempfile.TemporaryDirectory() as folder:
            source, destination = Path(folder) / "source.csv", Path(folder) / "reviewed.csv"
            source.write_text('id,note\n1,"line one\nline two"\n2,ok\n', encoding="utf-8")
            apply_operations(source, destination, [])
            with destination.open(encoding="utf-8", newline="") as handle:
                rows = list(csv.DictReader(handle))
        self.assertEqual(rows[0]["note"], "line one\nline two")

    def test_chained_corrections_are_ordered_by_dependency(self) -> None:
        with tempfile.TemporaryDirectory() as folder:
            source, destination = Path(folder) / "source.csv", Path(folder) / "reviewed.csv"
            source.write_text("id,name\n1, david lee \n", encoding="utf-8")
            operations = [
                {"type": "replace", "column": "name", "before": "david lee", "after": "David Lee", "rows": [1]},
                {"type": "trim", "column": "name", "before": " david lee ", "after": "david lee", "rows": [1]},
            ]
            apply_operations(source, destination, operations)
            self.assertIn("1,David Lee\n", destination.read_text(encoding="utf-8"))

    def test_generated_r_script_is_traceable(self) -> None:
        operation = {"type": "trim", "column": "occupation", "before": "Teacher  ", "after": "Teacher", "rows": [4]}
        script = generate_r_script("source.csv", "reviewed.csv", [operation], ",")
        self.assertIn("accepted trim correction", script)
        self.assertIn('"occupation"', script)
        self.assertIn("target_rows <- c(4)", script)
        self.assertIn("cleaned/reviewed.csv", script)
        self.assertIn("col_character", script)
        self.assertIn("na = character()", script)
        self.assertIn("length(matching_rows) == length(target_rows)", script)

    def test_change_ledger_rejects_an_unapproved_shape_preserving_edit(self) -> None:
        with tempfile.TemporaryDirectory() as folder:
            source = Path(folder) / "source.csv"
            reviewed = Path(folder) / "reviewed.csv"
            source.write_text("id,value\n001, dirty \n002,valid\n", encoding="utf-8")
            reviewed.write_text("id,value\n001,dirty\n002,corrupted\n", encoding="utf-8")
            operation = {"type": "trim", "column": "value", "before": " dirty ", "after": "dirty", "rows": [1]}
            item = {
                "id": "file_fixture",
                "format": "csv",
                "profile": {"encoding": "utf-8", "delimiter": ","},
            }
            with self.assertRaisesRegex(ValueError, "does not match the approved operation plan"):
                validate_operation_result(item, source, reviewed, [operation])

    def test_change_ledger_preserves_leading_zero_identifiers(self) -> None:
        with tempfile.TemporaryDirectory() as folder:
            source = Path(folder) / "source.csv"
            reviewed = Path(folder) / "reviewed.csv"
            source.write_text("id,value\n001, dirty \n002,valid\n", encoding="utf-8")
            operation = {"type": "trim", "column": "value", "before": " dirty ", "after": "dirty", "rows": [1]}
            apply_operations(source, reviewed, [operation])
            item = {
                "id": "file_fixture",
                "format": "csv",
                "profile": {"encoding": "utf-8", "delimiter": ","},
            }
            validation = validate_operation_result(item, source, reviewed, [operation])
        self.assertEqual(validation["expected_changed_cells"], 1)
        self.assertEqual(validation["unexpected_changed_cells"], 0)

    def test_deterministic_cell_cleaning_is_idempotent(self) -> None:
        with tempfile.TemporaryDirectory() as folder:
            source = Path(folder) / "source.csv"
            first = Path(folder) / "first.csv"
            second = Path(folder) / "second.csv"
            source.write_text("id,value\n001, dirty \n", encoding="utf-8")
            operation = {"type": "trim", "column": "value", "before": " dirty ", "after": "dirty", "rows": [1]}
            apply_operations(source, first, [operation])
            apply_operations(first, second, [operation])
            self.assertEqual(first.read_bytes(), second.read_bytes())

    def test_generated_r_script_applies_cell_changes_before_deletions_and_type_conversion(self) -> None:
        operations = [
            {"type": "parse_type", "column": "age", "before": "forty", "after": 40, "rows": [2], "expected": "integer"},
            {"type": "delete_rows", "rows": [3]},
            {"type": "normalize_missing", "column": "age", "before": "nan", "after": "", "rows": [4]},
        ]
        script = generate_r_script("source.csv", "reviewed.csv", operations, ",")
        self.assertLess(script.index("normalize_missing correction"), script.index("parse_integer"))
        self.assertLess(script.index("parse_integer"), script.index("Remove explicitly approved duplicate rows"))
        self.assertIn("matching_rows", script)

    def test_profiles_xlsx_and_keeps_sheet_metadata(self) -> None:
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as folder:
            path = Path(folder) / "study.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Participants"
            sheet.append(["participant_id", "occupation", "age"])
            sheet.append(["P001", "Teacher  ", 30])
            sheet.append(["P002", "Nurse", 29])
            workbook.create_sheet("Codebook")
            workbook.save(path)
            result = profile_dataset(path, path.name)
        self.assertEqual(result["table_name"], "Participants")
        self.assertEqual(result["row_count"], 2)
        self.assertEqual(result["format_metadata"]["sheets"], ["Participants", "Codebook"])
        self.assertIn("whitespace", {item["category"] for item in result["findings"]})

    def test_xlsx_export_preserves_other_sheets(self) -> None:
        from openpyxl import Workbook, load_workbook

        with tempfile.TemporaryDirectory() as folder:
            source, output = Path(folder) / "source.xlsx", Path(folder) / "reviewed.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Participants"
            sheet.append(["participant_id", "occupation"])
            sheet.append(["P001", "Teacher  "])
            workbook.create_sheet("Codebook")["A1"] = "Do not change"
            workbook.save(source)
            operation = {"type": "trim", "column": "occupation", "before": "Teacher  ", "after": "Teacher", "rows": [1]}
            profile = profile_dataset(source, source.name)
            apply_xlsx_operations(source, output, [operation], "Participants")
            reviewed = load_workbook(output)
            original = load_workbook(source)
            self.assertEqual(reviewed["Participants"]["B2"].value, "Teacher")
            self.assertEqual(reviewed["Codebook"]["A1"].value, "Do not change")
            self.assertEqual(original["Participants"]["B2"].value, "Teacher  ")
            item = {"id": "file_xlsx", "format": "xlsx", "profile": profile}
            self.assertEqual(validate_operation_result(item, source, output, [operation])["status"], "passed")
            self.assertEqual(validate_format_contract(item, source, output, [operation])["status"], "passed")

    def test_format_specific_r_scripts(self) -> None:
        script = generate_format_r_script("study.sav", "study_reviewed.sav", [], "sav")
        self.assertIn("haven::read_sav", script)
        self.assertIn("haven::write_sav", script)

    def test_statistical_formats_round_trip_and_preserve_original(self) -> None:
        import pandas as pd
        import pyreadr
        import pyreadstat

        frame = pd.DataFrame({"participant_id": ["P001", "P002"], "occupation": ["Teacher  ", "Nurse"], "score": [1.0, 2.0]})
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            sources = {"sav": root / "study.sav", "dta": root / "study.dta", "rds": root / "study.rds"}
            pyreadstat.write_sav(frame, sources["sav"], column_labels={"score": "Study score"})
            pyreadstat.write_dta(frame, sources["dta"], column_labels={"score": "Study score"})
            pyreadr.write_rds(sources["rds"], frame)
            for file_format, source in sources.items():
                with self.subTest(file_format=file_format):
                    before = source.read_bytes()
                    profile = profile_dataset(source, source.name)
                    self.assertEqual(profile["row_count"], 2)
                    destination = root / f"reviewed.{file_format}"
                    operation = {"type": "replace", "column": "score", "before": 1.0, "after": 3.0, "rows": [1]}
                    from backend.app.exporting import apply_statistical_operations
                    apply_statistical_operations(source, destination, [operation], file_format)
                    reviewed = profile_dataset(destination, destination.name)
                    item = {"id": f"file_{file_format}", "format": file_format, "profile": profile}
                    self.assertEqual(validate_operation_result(item, source, destination, [operation])["status"], "passed")
                    self.assertEqual(validate_format_contract(item, source, destination, [operation])["status"], "passed")
                    self.assertEqual(reviewed["row_count"], 2)
                    self.assertEqual(source.read_bytes(), before)


if __name__ == "__main__":
    unittest.main()
